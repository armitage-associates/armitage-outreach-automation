import argparse
import asyncio
import json
import logging
from datetime import datetime
from pathlib import Path
from scraper import scrape_companies, read_companies_from_csv
from salesforce import import_companies_from_salesforce, push_to_salesforce
from utils.email_client import send_all_reports, send_owner_digests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger(__name__)

def run(
    recipients: list[str] = None,
    send_digest: bool = True,
    company: str = None,
    scrape_only: bool = False,
    deliver_only: bool = False,
    to_excel: bool = False,
    skip_analysis: bool = False,
    batch: str = None,
    limit: int = None,
    quarter: str = None,
):
    """
    Run the full scraping and email pipeline.

    Args:
        recipients: List of email addresses to send reports to.
        send_digest: If True, send one digest email. If False, send individual emails per company.
        company: If provided, only process this single company.
        scrape_only: If True, import + scrape only — skip push, email, cleanup.
        deliver_only: If True, push + email + cleanup only — skip import + scrape.
        to_excel: If True, write results to Excel instead of pushing to Salesforce.
        skip_analysis: If True, skip OpenAI analysis and contact scraping.
        batch: Batch spec like "1/4" meaning "batch 1 of 4".
        limit: If provided, only process the first N companies from the list.
    """
    # ── Scrape phase ──
    if not deliver_only and not to_excel:
        if company:
            logger.info(f"Single-company mode: {company}")
            import_companies_from_salesforce()
            companies = read_companies_from_csv()
            match = [(name, loc) for name, loc in companies if name.lower() == company.lower()]
            if not match:
                logger.error(f"Company '{company}' not found in companies.csv")
                return
            logger.info(f"Found: {match[0][0]} in {match[0][1]}")
            asyncio.run(scrape_companies(match, inter_delay=False, skip_analysis=skip_analysis))
        elif batch:
            batch_num, total_batches = _parse_batch(batch)
            if not scrape_only:
                import_companies_from_salesforce()
            companies = read_companies_from_csv()
            if limit:
                companies = companies[:limit]
                logger.info(f"Limited to first {limit} companies")
            chunk = _get_batch_slice(companies, batch_num, total_batches)
            logger.info(f"Batch {batch_num}/{total_batches}: processing {len(chunk)} of {len(companies)} companies")
            for name, loc in chunk:
                logger.info(f"  - {name}")
            asyncio.run(scrape_companies(chunk, skip_analysis=skip_analysis))
        else:
            import_companies_from_salesforce()
            companies = read_companies_from_csv()
            if limit:
                companies = companies[:limit]
                logger.info(f"Limited to first {limit} companies")
            asyncio.run(scrape_companies(companies, skip_analysis=skip_analysis))

    if scrape_only:
        logger.info("Scrape-only mode: skipping push, email, and cleanup")
        return

    # ── Deliver phase ──
    if to_excel:
        write_results_to_excel(quarter_override=quarter)
        cleanup()
        return

    push_to_salesforce()

    if send_digest:
        send_owner_digests(fallback_recipients=recipients)
    elif recipients:
        send_all_reports(recipients)
    else:
        logger.warning("No recipients configured, pass recipients to run().")

    cleanup()


def _parse_batch(batch_str: str) -> tuple[int, int]:
    """Parse '1/4' into (1, 4)."""
    parts = batch_str.split("/")
    if len(parts) != 2:
        raise ValueError(f"Invalid batch format '{batch_str}', expected 'N/M' (e.g. '1/4')")
    batch_num, total = int(parts[0]), int(parts[1])
    if batch_num < 1 or batch_num > total:
        raise ValueError(f"Batch number must be between 1 and {total}, got {batch_num}")
    return batch_num, total


def _get_batch_slice(companies: list, batch_num: int, total_batches: int, batch_size: int = 5) -> list:
    """Return a fixed-size slice of companies for the given batch number (1-indexed).

    Each batch contains exactly `batch_size` companies, except the last batch
    which may contain fewer.
    """
    start = (batch_num - 1) * batch_size
    end = start + batch_size
    return companies[start:end]


def write_results_to_excel(quarter_override=None):
    """Read output JSONs and append news + LinkedIn posts to a sheet in GOWT_mid_low.xlsx.

    Creates the sheet if it doesn't exist; appends rows if it does (so Medium and Low
    workflows can write to the same quarter sheet sequentially).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    project_root = Path(__file__).parent
    excel_path = project_root / "GOWT_mid_low.xlsx"
    output_dir = project_root / "data" / "output"

    if quarter_override:
        quarter = quarter_override
    else:
        now = datetime.now()
        quarter = f"Q{(now.month - 1) // 3 + 1} {now.year}"

    results = []
    for json_file in sorted(output_dir.glob("*.json")):
        try:
            with open(json_file) as f:
                data = json.load(f)
            if "company" in data:
                results.append(data)
        except Exception as e:
            logger.warning(f"Could not read {json_file}: {e}")

    if not results:
        logger.warning("No output files found for Excel export")
        return

    wb = load_workbook(excel_path)
    sheet_name = f"{quarter} News"

    headers = ["Company", "Location", "News Articles", "LinkedIn Posts", "LinkedIn URL"]
    header_font = Font(bold=True, size=11, color="2F5496")
    header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    for name in wb.sheetnames:
        if name.endswith(" News") and name != sheet_name:
            wb.remove(wb[name])
            logger.info(f"Removed previous quarter sheet '{name}'")

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        start_row = ws.max_row + 1
        logger.info(f"Appending to existing sheet '{sheet_name}' from row {start_row}")
    else:
        ws = wb.create_sheet(sheet_name)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 80
        ws.column_dimensions['E'].width = 50
        start_row = 2

    for row_idx, data in enumerate(results, start_row):
        ws.cell(row=row_idx, column=1, value=data.get("company", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=data.get("location", "")).border = thin_border

        articles = data.get("articles", [])
        article_lines = []
        for a in articles:
            headline = a.get("headline", "")
            date = a.get("date", "")
            summary = a.get("summary", "")
            line = f"{headline} ({date})"
            if summary:
                line += f"\n{summary}"
            article_lines.append(line)
        cell = ws.cell(row=row_idx, column=3, value="\n\n".join(article_lines))
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        posts = data.get("posts", [])
        post_lines = []
        for p in posts:
            content = p.get("content", p.get("summary", ""))
            date = p.get("date", "")
            if len(content) > 500:
                content = content[:500] + "..."
            post_lines.append(f"[{date}] {content}")
        cell = ws.cell(row=row_idx, column=4, value="\n\n".join(post_lines))
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        ws.cell(row=row_idx, column=5, value=data.get("linkedin_url", "")).border = thin_border

    wb.save(excel_path)
    logger.info(f"Wrote {len(results)} companies to sheet '{sheet_name}' in {excel_path} (starting row {start_row})")


def cleanup(input_dir: str = "data/input", output_dir: str = "data/output"):
    """Delete all files from data/input and data/output directories."""
    base = Path(__file__).parent
    for dir_path in (base / input_dir, base / output_dir):
        if not dir_path.exists():
            continue
        for file in dir_path.iterdir():
            if file.is_file():
                file.unlink()
                logger.info(f"Deleted {file}")
    logger.info("Cleanup complete")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Armitage automation pipeline")
    parser.add_argument(
        "--company",
        type=str,
        help="Run pipeline for a single company (must match a name in companies.csv)",
    )
    parser.add_argument(
        "--batch",
        type=str,
        help="Process a batch of companies, e.g. '1/4' for batch 1 of 4",
    )
    parser.add_argument(
        "--scrape-only",
        action="store_true",
        help="Import and scrape only — skip push, email, and cleanup. Output stays in data/output/",
    )
    parser.add_argument(
        "--deliver-only",
        action="store_true",
        help="Push + email + cleanup only — use existing output files, skip scraping",
    )
    parser.add_argument(
        "--to-excel",
        action="store_true",
        help="Write results to Excel (GOWT_mid_low.xlsx) instead of pushing to Salesforce. Reads existing output files.",
    )
    parser.add_argument(
        "--skip-analysis",
        action="store_true",
        help="Skip OpenAI analysis and contact scraping. Raw news + LinkedIn posts only.",
    )
    parser.add_argument(
        "--no-email",
        action="store_true",
        help="Skip sending emails (useful for testing)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Only process the first N companies (useful for testing)",
    )
    parser.add_argument(
        "--quarter",
        type=str,
        help="Override quarter label for --to-excel, e.g. 'Q3 2026'",
    )
    args = parser.parse_args()

    if args.scrape_only and args.deliver_only:
        parser.error("Cannot use --scrape-only and --deliver-only together")

    if args.no_email:
        run(
            company=args.company,
            send_digest=False,
            scrape_only=args.scrape_only,
            deliver_only=args.deliver_only,
            to_excel=args.to_excel,
            skip_analysis=args.skip_analysis,
            batch=args.batch,
            limit=args.limit,
            quarter=args.quarter,
        )
    else:
        run(
            ["mwan0165@student.monash.edu"],
            company=args.company,
            scrape_only=args.scrape_only,
            deliver_only=args.deliver_only,
            to_excel=args.to_excel,
            skip_analysis=args.skip_analysis,
            batch=args.batch,
            limit=args.limit,
            quarter=args.quarter,
        )
