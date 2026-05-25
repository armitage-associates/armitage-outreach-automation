"""
Scrape LinkedIn employee counts for GOWT Medium/Low companies and update the FTE tracking Excel.

Usage:
    python salesforce/fte_scrape.py                  # auto-detect current quarter
    python salesforce/fte_scrape.py --quarter "Q3 2026"
    python salesforce/fte_scrape.py --dry-run
"""

import argparse
import csv
import logging
import os
import sys
import time
from datetime import datetime

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

DATASET_ID = "gd_l1vikfnt1wgvvqz95w"
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "GOWT_mid_low.xlsx")
LINKEDIN_MAP_PATH = os.path.join(PROJECT_ROOT, "data", "fte_company_linkedin_map.csv")
SHEET_NAME = "FTE Tracking"

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
QUARTER_FILL = PatternFill(start_color="EBF1F8", end_color="EBF1F8", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="2F5496")


def get_current_quarter():
    now = datetime.now()
    q = (now.month - 1) // 3 + 1
    return f"Q{q} {now.year}"


def refresh_owners(ws):
    """Refresh the Owner column (C) from Salesforce opportunity owners."""
    import urllib.parse
    sys.path.insert(0, PROJECT_ROOT)
    from salesforce import get_access_token, sf_get

    token = get_access_token()
    soql = "SELECT Name, Owner.Name FROM Opportunity WHERE GOWT_Priority__c IN ('Medium', 'Low') ORDER BY Name"
    result = sf_get(f"query/?q={urllib.parse.quote(soql)}", token)

    owner_map = {}
    for r in result.get("records", []):
        name = r.get("Name", "").strip()
        owner = r.get("Owner", {}).get("Name", "")
        if name:
            owner_map[name] = owner

    while not result.get("done", True) and "nextRecordsUrl" in result:
        next_url = result["nextRecordsUrl"].split("/services/data/v62.0/")[-1]
        result = sf_get(next_url, token)
        for r in result.get("records", []):
            name = r.get("Name", "").strip()
            owner = r.get("Owner", {}).get("Name", "")
            if name:
                owner_map[name] = owner

    updated = 0
    for row in range(2, ws.max_row + 1):
        company = ws.cell(row=row, column=1).value
        if company and company.strip() in owner_map:
            ws.cell(row=row, column=3, value=owner_map[company.strip()])
            updated += 1

    logger.info(f"Refreshed owners for {updated}/{ws.max_row - 1} companies")
    return updated


def load_companies():
    companies = []
    with open(LINKEDIN_MAP_PATH, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("linkedin_slug"):
                companies.append((row["company_name"], row["linkedin_slug"]))
    return companies


def scrape_employee_counts(companies, api_key):
    results = {}
    batch_size = 50
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    total_batches = (len(companies) + batch_size - 1) // batch_size

    for batch_start in range(0, len(companies), batch_size):
        batch = companies[batch_start:batch_start + batch_size]
        batch_num = batch_start // batch_size + 1

        inputs = [{"url": f"https://www.linkedin.com/company/{slug}"} for _, slug in batch]

        trigger_url = f"https://api.brightdata.com/datasets/v3/trigger?dataset_id={DATASET_ID}&format=json&notify=false"
        resp = requests.post(trigger_url, headers=headers, json=inputs, timeout=30)

        if not resp.ok:
            logger.error(f"Batch {batch_num}/{total_batches}: trigger failed ({resp.status_code})")
            continue

        snapshot_id = resp.json().get("snapshot_id")
        if not snapshot_id:
            logger.error(f"Batch {batch_num}/{total_batches}: no snapshot_id")
            continue

        poll_url = f"https://api.brightdata.com/datasets/v3/progress/{snapshot_id}"
        max_wait = 600
        elapsed = 0
        status = None

        while elapsed < max_wait:
            time.sleep(15)
            elapsed += 15
            try:
                progress_resp = requests.get(poll_url, headers=headers)
                if progress_resp.ok:
                    status = progress_resp.json().get("status")
                    if status in ("ready", "failed"):
                        break
            except Exception:
                continue

        if status != "ready":
            logger.error(f"Batch {batch_num}/{total_batches}: {'timeout' if status != 'failed' else 'failed'}")
            continue

        download_url = f"https://api.brightdata.com/datasets/v3/snapshot/{snapshot_id}?format=json"
        dl_resp = requests.get(download_url, headers=headers, timeout=60)

        if dl_resp.ok:
            data = dl_resp.json()
            if not isinstance(data, list):
                logger.warning(f"Batch {batch_num}/{total_batches}: unexpected response format, skipping")
                continue
            count = 0
            for record in data:
                if not isinstance(record, dict):
                    continue
                slug = record.get("id", "")
                if slug:
                    results[slug] = record.get("employees_in_linkedin")
                    count += 1
            logger.info(f"Batch {batch_num}/{total_batches}: {count} results ({elapsed}s)")
        else:
            logger.error(f"Batch {batch_num}/{total_batches}: download failed ({dl_resp.status_code})")

    return results


def find_change_columns(ws):
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h and "Change" in str(h) and "%" not in str(h):
            return col
    return None


def update_excel(results, quarter_label, dry_run=False):
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # Refresh owner column from Salesforce
    try:
        refresh_owners(ws)
    except Exception as e:
        logger.warning(f"Owner refresh failed (non-fatal): {e}")

    # Check if this quarter already has a column
    existing_col = None
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h and str(h).startswith(quarter_label):
            existing_col = col
            break

    if existing_col:
        quarter_col = existing_col
        logger.info(f"Overwriting existing column {quarter_col} for {quarter_label}")
    else:
        # Insert a new column before the Change columns
        change_col = find_change_columns(ws)
        if change_col:
            ws.insert_cols(change_col)
            quarter_col = change_col
        else:
            quarter_col = ws.max_column + 1

        # Style the header
        cell = ws.cell(row=1, column=quarter_col, value=f"{quarter_label} (LinkedIn)")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER
        logger.info(f"Inserted new column {quarter_col} for {quarter_label}")

    # Fill data — match by LinkedIn URL in column 7
    filled = 0
    for row in range(2, ws.max_row + 1):
        li_url = ws.cell(row=row, column=7).value
        if not li_url or "linkedin.com/company/" not in str(li_url):
            continue
        slug = str(li_url).split("linkedin.com/company/")[-1].strip("/").split("/")[0]
        employee_count = results.get(slug)

        if employee_count is not None:
            cell = ws.cell(row=row, column=quarter_col, value=employee_count)
            cell.number_format = '#,##0'
            cell.fill = QUARTER_FILL
            cell.border = THIN_BORDER
            filled += 1

    # Find previous quarter column (the one immediately before current quarter)
    # Column layout: A=Company, B=Location, C=Owner, D=Priority, E=Industry, F=Website, G=LinkedIn URL, H=Baseline FTE, I+=quarters
    prev_col = None
    prev_label = None
    if quarter_col > 9:
        prev_col = quarter_col - 1
        prev_label = ws.cell(row=1, column=prev_col).value
    elif quarter_col == 9:
        prev_col = 8
        prev_label = "Baseline FTE (SF)"

    # Update Change columns (now shifted if we inserted)
    change_col = find_change_columns(ws)
    if change_col and prev_col:
        pct_col = change_col + 1
        prev_short = str(prev_label).replace(" (LinkedIn)", "").replace(" (SF)", "") if prev_label else "prev"
        ws.cell(row=1, column=change_col, value=f"Change ({prev_short} vs {quarter_label})")

        for row in range(2, ws.max_row + 1):
            prev_fte = ws.cell(row=row, column=prev_col).value
            li_fte = ws.cell(row=row, column=quarter_col).value

            if prev_fte is not None and li_fte is not None:
                diff = li_fte - prev_fte
                cell = ws.cell(row=row, column=change_col, value=diff)
                cell.border = THIN_BORDER
                cell.number_format = '+#,##0;-#,##0;0'
                cell.font = Font(color="006100" if diff > 0 else "9C0006" if diff < 0 else "000000")

                if prev_fte > 0:
                    pct = (li_fte - prev_fte) / prev_fte
                    cell = ws.cell(row=row, column=pct_col, value=pct)
                    cell.border = THIN_BORDER
                    cell.number_format = '+0.0%;-0.0%;0.0%'
                    cell.font = Font(color="006100" if pct > 0 else "9C0006" if pct < 0 else "000000")
            else:
                ws.cell(row=row, column=change_col, value='').border = THIN_BORDER
                ws.cell(row=row, column=pct_col, value='').border = THIN_BORDER

    # Update summary sheet
    if "Summary" in wb.sheetnames:
        ws2 = wb["Summary"]
        ws2.cell(row=8, column=2, value=f"{quarter_label} (updated {datetime.now().strftime('%Y-%m-%d')})")

    logger.info(f"Filled {filled} companies for {quarter_label}")

    if not dry_run:
        wb.save(EXCEL_PATH)
        logger.info(f"Saved to {EXCEL_PATH}")
    else:
        logger.info("Dry run — not saving")

    return filled


def save_results_cache(results, quarter):
    import json
    cache_dir = os.path.join(PROJECT_ROOT, "data")
    os.makedirs(cache_dir, exist_ok=True)
    filename = f"fte_scrape_{quarter.replace(' ', '_')}.json"
    path = os.path.join(cache_dir, filename)
    with open(path, "w") as f:
        json.dump(results, f, indent=2)
    logger.info(f"Cached scrape results to {path}")
    return path


def load_results_cache(path):
    import json
    with open(path) as f:
        return json.load(f)


def main():
    parser = argparse.ArgumentParser(description="Scrape LinkedIn FTE counts and update Excel")
    parser.add_argument("--quarter", type=str, help="Quarter label, e.g. 'Q3 2026'. Auto-detected if omitted.")
    parser.add_argument("--dry-run", action="store_true", help="Scrape but don't write to Excel")
    parser.add_argument("--from-cache", type=str, help="Skip scraping, load results from a cached JSON file")
    parser.add_argument("--companies", type=str, help="Comma-separated company names to filter (e.g. 'Hardcat Pty Ltd,GreenBe')")
    args = parser.parse_args()

    quarter = args.quarter or get_current_quarter()
    logger.info(f"FTE scrape for {quarter}")

    if args.from_cache:
        results = load_results_cache(args.from_cache)
        logger.info(f"Loaded {len(results)} results from cache: {args.from_cache}")
    else:
        api_key = os.environ.get("BRIGHTDATA_API_KEY")
        if not api_key:
            logger.error("BRIGHTDATA_API_KEY not set")
            sys.exit(1)

        companies = load_companies()
        if args.companies:
            filter_names = {n.strip().lower() for n in args.companies.split(",")}
            companies = [(name, slug) for name, slug in companies if name.lower() in filter_names]
        logger.info(f"Loaded {len(companies)} companies from {LINKEDIN_MAP_PATH}")

        results = scrape_employee_counts(companies, api_key)
        logger.info(f"Scraped {len(results)} employee counts")

        save_results_cache(results, quarter)

    filled = update_excel(results, quarter, dry_run=args.dry_run)
    logger.info(f"Done: {filled}/{len(results)} companies updated for {quarter}")


if __name__ == "__main__":
    main()
