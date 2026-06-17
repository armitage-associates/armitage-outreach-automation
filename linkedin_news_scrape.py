"""
LinkedIn news scrape for GOWT High (monthly) and Medium (quarterly) companies.

Scrapes LinkedIn company posts via BrightData API and writes results to Excel.
No Perplexity, no OpenAI, no Salesforce push.

Usage:
    # High companies (monthly) — full run
    python linkedin_news_scrape.py --priority high

    # Medium companies (quarterly) — multi-job workflow
    python linkedin_news_scrape.py --priority medium --import-only --slice 1/5
    python linkedin_news_scrape.py --priority medium --scrape-only --batch 1/11
    python linkedin_news_scrape.py --priority medium --deliver-only --quarter "Q3 2026"

    # Testing
    python linkedin_news_scrape.py --priority high --limit 2 --dry-run
"""

import argparse
import csv
import json
import logging
import os
import re
import time
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from salesforce import get_access_token, sf_get

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent
HIGH_EXCEL = PROJECT_ROOT / "GOWT_high.xlsx"
MEDIUM_EXCEL = PROJECT_ROOT / "GOWT_mid_low.xlsx"
INPUT_DIR = PROJECT_ROOT / "data" / "input"
OUTPUT_DIR = PROJECT_ROOT / "data" / "output"
BRIGHTDATA_DATASET = "gd_lyy3tktm25m4avu764"

GOWT_SOQL = (
    "SELECT Name, fid5__c FROM Opportunity "
    "WHERE StageName = '8. Good opportunity wrong timing' "
    "AND GOWT_Priority__c = '{priority}' "
    "AND Transaction_type__c != '8. Portfolio company bolt-on' "
    "ORDER BY Name"
)

HEADER_FONT = Font(bold=True, size=11, color="2F5496")
HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


# ---------------------------------------------------------------------------
# Salesforce import
# ---------------------------------------------------------------------------

def import_gowt_companies(priority, token):
    """Query Salesforce for GOWT High or Medium companies.

    Returns list of (company_name, location) tuples.
    """
    soql = GOWT_SOQL.format(priority=priority.capitalize())
    result = sf_get("query/?q=" + urllib.parse.quote(soql), token)

    companies = []
    while True:
        for rec in result.get("records", []):
            companies.append((rec["Name"], rec.get("fid5__c", "")))
        next_url = result.get("nextRecordsUrl")
        if not next_url:
            break
        result = sf_get(next_url.split("/services/data/v62.0/")[-1], token)

    logger.info(f"Imported {len(companies)} GOWT {priority.capitalize()} companies from Salesforce")
    return companies


def slice_companies(companies, slice_spec):
    """Slice companies for multi-day runs. slice_spec is '1/5' meaning slice 1 of 5."""
    slice_num, total_slices = map(int, slice_spec.split("/"))
    chunk_size = len(companies) // total_slices
    remainder = len(companies) % total_slices
    start = 0
    for i in range(1, slice_num):
        start += chunk_size + (1 if i <= remainder else 0)
    size = chunk_size + (1 if slice_num <= remainder else 0)
    result = companies[start:start + size]
    logger.info(f"Slice {slice_num}/{total_slices}: {len(result)} companies (of {len(companies)} total)")
    return result


def get_batch_slice(companies, batch_spec, batch_size=5):
    """Return a batch slice. batch_spec is '1/10' meaning batch 1 of 10."""
    parts = list(map(int, batch_spec.split("/")))
    start = (parts[0] - 1) * batch_size
    end = start + batch_size
    return companies[start:end]


# ---------------------------------------------------------------------------
# LinkedIn slug resolution via BrightData SERP
# ---------------------------------------------------------------------------

def resolve_linkedin_slug(company_name, location):
    """Use BrightData SERP to find a company's LinkedIn slug.

    Returns the slug (e.g. 'urbanx-technology') or None.
    """
    api_key = os.getenv("BRIGHTDATA_API_KEY")
    if not api_key:
        logger.error("BRIGHTDATA_API_KEY not set")
        return None

    query = f"{company_name} {location} site:linkedin.com/company"
    payload = {
        "zone": "serp_api1",
        "url": f"https://www.google.com/search?q={urllib.parse.quote_plus(query)}&gl=au&hl=en&brd_json=1",
        "format": "raw",
    }

    try:
        resp = requests.post(
            "https://api.brightdata.com/request",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=30,
        )
        if not resp.ok:
            logger.warning(f"SERP API error for {company_name}: {resp.status_code}")
            return None

        organic = resp.json().get("organic", [])
        for result in organic:
            link = result.get("link", "")
            match = re.search(r"linkedin\.com/company/([^/?#]+)", link)
            if match:
                slug = match.group(1).rstrip("/")
                logger.info(f"Resolved LinkedIn slug for {company_name}: {slug}")
                return slug

        logger.warning(f"No LinkedIn company page found for {company_name}")
        return None
    except Exception as e:
        logger.error(f"SERP lookup failed for {company_name}: {e}")
        return None


def resolve_slugs_for_companies(companies):
    """Resolve LinkedIn slugs for a list of (name, location) tuples.

    Returns list of (name, location, linkedin_slug) tuples.
    """
    results = []
    for name, location in companies:
        slug = resolve_linkedin_slug(name, location)
        if slug:
            results.append((name, location, slug))
        else:
            logger.warning(f"Skipping {name} — no LinkedIn slug found")
    logger.info(f"Resolved {len(results)}/{len(companies)} LinkedIn slugs")
    return results


# ---------------------------------------------------------------------------
# LinkedIn posts scraping via BrightData API
# ---------------------------------------------------------------------------

def scrape_linkedin_posts(company_name, linkedin_slug, start_date, end_date):
    """Scrape LinkedIn company posts via BrightData API.

    Args:
        company_name: Company name (for logging)
        linkedin_slug: LinkedIn company slug
        start_date: Start date (datetime)
        end_date: End date (datetime)

    Returns list of post dicts [{title, post_text, date_posted}] or None on failure.
    """
    api_key = os.getenv("BRIGHTDATA_API_KEY")
    if not api_key:
        logger.error("BRIGHTDATA_API_KEY not set")
        return None

    company_url = f"https://www.linkedin.com/company/{linkedin_slug}"
    start_str = start_date.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    end_str = end_date.strftime("%Y-%m-%dT%H:%M:%S.000Z")

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    logger.info(f"Scraping LinkedIn posts for {company_name} ({start_str[:10]} to {end_str[:10]})")

    try:
        # Trigger async scrape
        resp = requests.post(
            f"https://api.brightdata.com/datasets/v3/trigger"
            f"?dataset_id={BRIGHTDATA_DATASET}"
            f"&custom_output_fields=title%2Cpost_text%2Cdate_posted"
            f"&notify=false&type=discover_new&discover_by=company_url",
            headers=headers,
            data=json.dumps({
                "input": [{
                    "url": company_url,
                    "start_date": start_str,
                    "end_date": end_str,
                }],
            }),
        )

        if not resp.ok:
            logger.error(f"Trigger failed for {company_name} ({resp.status_code}): {resp.text[:300]}")
            return None

        snapshot_id = resp.json().get("snapshot_id")
        if not snapshot_id:
            logger.error(f"No snapshot_id for {company_name}: {resp.text[:300]}")
            return None

        logger.info(f"Triggered snapshot {snapshot_id} for {company_name}")

        # Poll for completion
        max_wait = 900
        poll_interval = 30
        elapsed = 0
        while elapsed < max_wait:
            time.sleep(poll_interval)
            elapsed += poll_interval
            progress = requests.get(
                f"https://api.brightdata.com/datasets/v3/progress/{snapshot_id}",
                headers={"Authorization": f"Bearer {api_key}"},
            )
            if not progress.ok:
                continue
            status = progress.json().get("status")
            logger.info(f"  {company_name} snapshot {snapshot_id}: {status} ({elapsed}s)")
            if status == "ready":
                break
            if status == "failed":
                logger.error(f"Snapshot failed for {company_name}")
                return None
        else:
            logger.error(f"Timeout waiting for {company_name} snapshot")
            return None

        # Download results
        download = requests.get(
            f"https://api.brightdata.com/datasets/v3/snapshot/{snapshot_id}?format=json",
            headers={"Authorization": f"Bearer {api_key}"},
        )
        if not download.ok:
            logger.error(f"Download failed for {company_name}: {download.status_code}")
            return None

        parsed = json.loads(download.text.strip())
        if isinstance(parsed, list):
            posts = [p for p in parsed if isinstance(p, dict) and "post_text" in p]
        elif isinstance(parsed, dict) and "post_text" in parsed:
            posts = [parsed]
        else:
            posts = []

        logger.info(f"Got {len(posts)} posts for {company_name}")
        return posts if posts else None

    except Exception as e:
        logger.error(f"LinkedIn scrape failed for {company_name}: {e}")
        return None


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def write_high_excel(results, month_label):
    """Write per-company tabs to GOWT_high.xlsx. Replaces entire workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    headers = ["Date Posted", "Title", "Post Text"]

    for company_name, location, linkedin_slug, posts in results:
        tab_name = company_name[:31]
        ws = wb.create_sheet(tab_name)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 100

        if not posts:
            ws.cell(row=2, column=1, value="No posts found").border = THIN_BORDER
            continue

        sorted_posts = sorted(posts, key=lambda p: p.get("date_posted", ""), reverse=True)
        for row_idx, post in enumerate(sorted_posts, 2):
            ws.cell(row=row_idx, column=1, value=post.get("date_posted", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=post.get("title", "")).border = THIN_BORDER
            cell = ws.cell(row=row_idx, column=3, value=post.get("post_text", ""))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = THIN_BORDER

    wb.save(HIGH_EXCEL)
    logger.info(f"Wrote {len(results)} company tabs to {HIGH_EXCEL} ({month_label})")


def write_medium_excel(results, quarter_label):
    """Write/append combined news tab to GOWT_mid_low.xlsx."""
    wb = load_workbook(MEDIUM_EXCEL)
    sheet_name = f"{quarter_label} News"

    # Remove previous News sheets
    for name in wb.sheetnames:
        if name.endswith(" News") and name != sheet_name:
            wb.remove(wb[name])
            logger.info(f"Removed previous sheet '{name}'")

    headers = ["Company", "Location", "LinkedIn Posts", "LinkedIn URL"]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        start_row = ws.max_row + 1
        logger.info(f"Appending to '{sheet_name}' from row {start_row}")
    else:
        ws = wb.create_sheet(sheet_name)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 50
        start_row = 2

    for row_idx, (company_name, location, linkedin_slug, posts) in enumerate(results, start_row):
        ws.cell(row=row_idx, column=1, value=company_name).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=location).border = THIN_BORDER

        post_lines = []
        for p in (posts or []):
            text = p.get("post_text", "")
            date = p.get("date_posted", "")
            if len(text) > 500:
                text = text[:500] + "..."
            post_lines.append(f"[{date}] {text}")
        cell = ws.cell(row=row_idx, column=3, value="\n\n".join(post_lines))
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = THIN_BORDER

        linkedin_url = f"https://www.linkedin.com/company/{linkedin_slug}" if linkedin_slug else ""
        ws.cell(row=row_idx, column=4, value=linkedin_url).border = THIN_BORDER

    wb.save(MEDIUM_EXCEL)
    logger.info(f"Wrote {len(results)} companies to '{sheet_name}' in {MEDIUM_EXCEL}")


# ---------------------------------------------------------------------------
# OneDrive upload
# ---------------------------------------------------------------------------

def upload_to_onedrive(file_path):
    """Upload file to OneDrive."""
    try:
        import onedrive
        token = onedrive.get_access_token()
        onedrive.upload(str(file_path), token)
    except Exception as e:
        logger.error(f"OneDrive upload failed: {e}")


# ---------------------------------------------------------------------------
# CSV I/O for multi-job workflow
# ---------------------------------------------------------------------------

def write_companies_csv(companies_with_slugs):
    """Write (name, location, slug) tuples to CSV for scrape jobs."""
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path = INPUT_DIR / "companies.csv"
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["company", "location", "linkedin_slug"])
        writer.writerows(companies_with_slugs)
    logger.info(f"Wrote {len(companies_with_slugs)} companies to {csv_path}")


def read_companies_csv():
    """Read companies CSV written by import phase."""
    csv_path = INPUT_DIR / "companies.csv"
    with open(csv_path) as f:
        reader = csv.DictReader(f)
        return [(r["company"], r["location"], r["linkedin_slug"]) for r in reader]


def write_output_json(company_name, location, linkedin_slug, posts):
    """Write scrape results for one company to JSON."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = {
        "company": company_name,
        "location": location,
        "linkedin_slug": linkedin_slug,
        "posts": posts or [],
    }
    path = OUTPUT_DIR / f"{company_name}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)


def read_output_jsons():
    """Read all output JSONs into results list."""
    results = []
    for json_file in sorted(OUTPUT_DIR.glob("*.json")):
        try:
            with open(json_file) as f:
                data = json.load(f)
            if "company" in data:
                results.append((
                    data["company"],
                    data.get("location", ""),
                    data.get("linkedin_slug", ""),
                    data.get("posts", []),
                ))
        except Exception as e:
            logger.warning(f"Could not read {json_file}: {e}")
    return results


def cleanup():
    """Delete all files from data/input and data/output."""
    for dir_path in (INPUT_DIR, OUTPUT_DIR):
        if not dir_path.exists():
            continue
        for f in dir_path.iterdir():
            if f.is_file():
                f.unlink()
    logger.info("Cleanup complete")


# ---------------------------------------------------------------------------
# Date range helpers
# ---------------------------------------------------------------------------

def get_month_date_range(month_override=None):
    """Return (start, end, label) for the past month."""
    now = datetime.now()
    if month_override:
        dt = datetime.strptime(month_override, "%b %Y")
        end = dt.replace(day=1) + timedelta(days=32)
        end = end.replace(day=1)
        start = dt.replace(day=1)
    else:
        end = now
        start = end - timedelta(days=30)
    label = month_override or now.strftime("%b %Y")
    return start, end, label


def get_quarter_date_range(quarter_override=None):
    """Return (start, end, label) for the past quarter."""
    now = datetime.now()
    if quarter_override:
        parts = quarter_override.split()
        q = int(parts[0][1])
        year = int(parts[1])
        quarter_start_month = (q - 1) * 3 + 1
        start = datetime(year, quarter_start_month, 1)
        if q == 4:
            end = datetime(year + 1, 1, 1)
        else:
            end = datetime(year, quarter_start_month + 3, 1)
    else:
        end = now
        start = end - timedelta(days=90)
    label = quarter_override or f"Q{(now.month - 1) // 3 + 1} {now.year}"
    return start, end, label


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(args):
    priority = args.priority.lower()
    is_high = priority == "high"

    if is_high:
        start_date, end_date, period_label = get_month_date_range(args.month)
    else:
        start_date, end_date, period_label = get_quarter_date_range(args.quarter)

    logger.info(f"=== GOWT {priority.capitalize()} scrape: {period_label} ===")
    logger.info(f"Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    # --- Deliver-only mode ---
    if args.deliver_only:
        results = read_output_jsons()
        if not results:
            logger.warning("No output files found")
            return
        if is_high:
            write_high_excel(results, period_label)
            upload_to_onedrive(HIGH_EXCEL)
        else:
            write_medium_excel(results, period_label)
            upload_to_onedrive(MEDIUM_EXCEL)
        cleanup()
        return

    # --- Import phase ---
    if not args.scrape_only:
        token = get_access_token()
        companies = import_gowt_companies(priority, token)

        if args.slice:
            companies = slice_companies(companies, args.slice)

        if args.limit:
            companies = companies[:args.limit]
            logger.info(f"Limited to {args.limit} companies")

        companies_with_slugs = resolve_slugs_for_companies(companies)

        if args.import_only:
            write_companies_csv(companies_with_slugs)
            return

    # --- Scrape phase ---
    if args.scrape_only:
        companies_with_slugs = read_companies_csv()
        if args.batch:
            companies_with_slugs = get_batch_slice(companies_with_slugs, args.batch)
        if args.limit:
            companies_with_slugs = companies_with_slugs[:args.limit]
            logger.info(f"Limited to {args.limit} companies")

    results = []
    for name, location, slug in companies_with_slugs:
        posts = scrape_linkedin_posts(name, slug, start_date, end_date)
        results.append((name, location, slug, posts))

        if args.scrape_only:
            write_output_json(name, location, slug, posts)

    if args.scrape_only:
        logger.info(f"Scrape-only complete: {len(results)} companies processed")
        return

    # --- Write Excel + upload ---
    if args.dry_run:
        logger.info(f"Dry run complete: {len(results)} companies scraped, not writing Excel")
        for name, loc, slug, posts in results:
            count = len(posts) if posts else 0
            logger.info(f"  {name}: {count} posts")
        return

    if is_high:
        write_high_excel(results, period_label)
        upload_to_onedrive(HIGH_EXCEL)
    else:
        write_medium_excel(results, period_label)
        upload_to_onedrive(MEDIUM_EXCEL)


def main():
    parser = argparse.ArgumentParser(description="LinkedIn news scrape for GOWT companies")
    parser.add_argument("--priority", required=True, choices=["high", "medium"],
                        help="GOWT priority level")
    parser.add_argument("--slice", type=str,
                        help="Slice spec for multi-day runs, e.g. '1/5'")
    parser.add_argument("--batch", type=str,
                        help="Batch spec for parallel jobs, e.g. '1/3'")
    parser.add_argument("--quarter", type=str,
                        help="Override quarter label, e.g. 'Q3 2026'")
    parser.add_argument("--month", type=str,
                        help="Override month label, e.g. 'Jun 2026'")
    parser.add_argument("--limit", type=int,
                        help="Only process first N companies (testing)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Scrape but don't write Excel")
    parser.add_argument("--import-only", action="store_true",
                        help="Only import from SF + resolve slugs, write CSV")
    parser.add_argument("--scrape-only", action="store_true",
                        help="Read CSV, scrape, write output JSONs")
    parser.add_argument("--deliver-only", action="store_true",
                        help="Read output JSONs, write Excel, upload OneDrive")
    args = parser.parse_args()

    modes = [args.import_only, args.scrape_only, args.deliver_only]
    if sum(modes) > 1:
        parser.error("Only one of --import-only, --scrape-only, --deliver-only can be used")

    run(args)


if __name__ == "__main__":
    main()
